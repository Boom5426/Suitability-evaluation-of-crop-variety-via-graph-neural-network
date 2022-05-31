import pandas as pd
import numpy as np
import h5py
from sklearn.metrics.pairwise import cosine_similarity as cos
from sklearn.metrics import pairwise_distances as pair
from sklearn.preprocessing import normalize
import torch
import torch.nn as nn
torch.set_default_tensor_type(torch.FloatTensor)
from torch.autograd import Variable
from pandas.core.frame import DataFrame
from networkx import karate_club_graph,to_numpy_matrix
import matplotlib.pyplot as plt
import networkx as nx
import torch as t
import torch.nn.functional as F
import scipy.sparse as sp
from dgl.nn.pytorch import GraphConv
import dgl
from torch.nn import Linear
import openpyxl
 
from openpyxl import load_workbook


#########################################################
######################   数据加载  ######################
#########################################################

data_path = './our_data.xls'
# save_path = './visual.xlsx'

wb = openpyxl.Workbook()
ws = wb.active

data =  pd.read_excel(data_path, sheet_name=0)
data_select = []
data_target = []
k = 0
a, b = 0, 0

for row in data.index.values:
    if data['评审状态'][row] == "淘汰":
        if a >= 5000: # 均衡采样
            continue
        data_target.append(0)
        a += 1
    elif data['评审状态'][row] == "续试" or data['评审状态'][row] == "续试1" :
        if b >= 5000:
            continue
        data_target.append(1)
        b += 1
    else:
        continue

    feature_1 = data['最高气温均值'][row]
    feature_2 = data['最高气温方差'][row]
    feature_3 = data['平均气温均值'][row]
    feature_4 = data['平均气温方差'][row]
    feature_5 = data['最低气温均值'][row]
    feature_6 = data['最低气温方差'][row]
    feature_7 = data['温差均值'][row]
    feature_8 = data['温差方差'][row]
    feature_9 = data['地面气压均值'][row]
    feature_10 = data['地面气压方差'][row]
    feature_11 = data['相对湿度均值'][row]
    feature_12 = data['相对湿度方差'][row]
    feature_13 = data['降水量均值'][row]
    feature_14 = data['降水量方差'][row]
    feature_15 = data['最大风速均值'][row]
    feature_16 = data['最大风速方差'][row]
    feature_17 = data['平均风速均值'][row]
    feature_18 = data['平均风速方差'][row]
    feature_19 = data['风向角度均值'][row]
    feature_20 = data['风向角度方差'][row]
    feature_21 = data['日照时间均值'][row]
    feature_22 = data['日照时间方差'][row]
    feature_23 = data['风力等级均值'][row]
    feature_24 = data['风力等级方差'][row]
    feature_25 = data['大斑病'][row] 
    feature_26 = data['倒伏率'][row]
    feature_27 = data['倒折率'][row]
    feature_28 = data['灰斑病'][row]
    feature_29 = data['株高'][row]
    if feature_29 < 10:
        feature_29 *= 100
    feature_30 = data['穗位高'][row]
    feature_31 = data['空杆率'][row]
    feature_32 = data['生育期'][row]
    feature_33 = data['穗腐病'][row]
    feature_34 = data['百粒重'][row]
    feature_35 = data['穗长'][row] 
    feature_36 = data['秃尖长'][row]
    feature_37 = data['鲜穗果重'][row]
    feature_38 = data['亩产'][row]
    feature_39 = data['增减产'][row]


    data_select.append([feature_1, feature_2, feature_3, feature_4, feature_5, feature_6, feature_7, \
        feature_8, feature_9, feature_10, feature_11, feature_12, feature_13, feature_14, feature_15, \
        feature_16, feature_17, feature_18, feature_19, feature_20, feature_21, feature_22, feature_23,\
        feature_24, feature_25, feature_26, feature_27, feature_28, feature_29, feature_30, \
        feature_31, feature_32, feature_33, feature_34, feature_35, feature_36, feature_37,  \
        feature_38, feature_39])


#########################################################
######################   数据处理  ######################
#########################################################
data_select = DataFrame(data_select)
# print(data_select.shape)

# 特征做标准化
numeric_features = data_select.dtypes[data_select.dtypes != 'object'].index
data_select[numeric_features] = data_select[numeric_features].apply(
    lambda x: (x - x.mean()) / (x.std()))

# for i in data_select:
#     feature_39 = i[-1]
#     print(feature_39)
#     ws.cell(row=i, column=5, value = feature_39)

# 标准化后，每个数值特征的均值变为0，所以可以直接用0来替换缺失值
# data_select[numeric_features] = data_select[numeric_features].fillna(0)

# dummy_na=True将缺失值也当作合法的特征值并为其创建指示特征
all_features = pd.get_dummies(data_select, dummy_na=True)
print(all_features.shape) # (198, 15)

# 确保没有NAN值
all_features = np.nan_to_num(all_features)

line = len(data_select)//5 * 4
# print(line) # 156

# train_data = data_select[:line]
# test_data = data_select[line:]

# n_train = train_data.shape[0]
# train_features = torch.tensor(all_features[500:2500], dtype=torch.float)
# train_labels = torch.tensor(data_target[500:2500], dtype=torch.float)

train_features = torch.tensor(all_features, dtype=torch.float)
train_labels = torch.tensor(data_target, dtype=torch.float)


# print(train_features[:3])
# print(train_labels[:20])

# 图网络模型

class GCN(nn.Module):
    def __init__(self, in_feats, hidden_size, num_classes):
        super(GCN, self).__init__()
        self.conv1 = GraphConv(in_feats,hidden_size)
        self.conv2 = GraphConv(hidden_size,hidden_size//2)
        self.conv3 = GraphConv(hidden_size//2,hidden_size//4)
        self.conv4 = GraphConv(hidden_size//4,hidden_size//8)
        self.conv5 = GraphConv(hidden_size//8,num_classes)
    def forward(self, g, x):
        # print(x.shape)
        # print(g.shape)
        x = torch.relu(self.conv1(g,x))
        x = torch.relu(self.conv2(g,x))
        x = torch.relu(self.conv3(g,x))
        x = torch.relu(self.conv4(g,x))
        x = torch.relu(self.conv5(g,x))

        return x

def load_graph():
    path = 'graph.txt'
    data = np.loadtxt('graph.txt')
    n, _ = data.shape
    # print(data.shape) # (107495, 2)

    idx = np.array([i for i in range(n)], dtype=np.int32)
    idx_map = {j: i for i, j in enumerate(idx)}
    edges_unordered = np.genfromtxt(path, dtype=np.int32)
    edges = np.array(list(map(idx_map.get, edges_unordered.flatten())),
                     dtype=np.int32).reshape(edges_unordered.shape)
    adj = sp.coo_matrix((np.ones(edges.shape[0]), (edges[:, 0], edges[:, 1])),
                        shape=(n, n), dtype=np.float32)

    # build symmetric adjacency matrix
    adj = adj + adj.T.multiply(adj.T > adj) - adj.multiply(adj.T > adj)
    adj = adj + sp.eye(adj.shape[0])
    adj = normalize(adj).todense()

    return adj

topk = 3
def construct_graph(features, label, method='heat'):
    # fname = 'graph.txt'
    num = len(label) # 要学习的标签数量
    dist = None

    # 计算特征之间的距离
    if method == 'heat':
        dist = -0.5 * pair(features) ** 2
        # 实现了对元素(可能是列表)的顺式访问
        dist = np.exp(dist) # e的dist次方
    elif method == 'cos': # 余弦距离
        features[features > 0] = 1
        dist = np.dot(features, features.T)
    elif method == 'ncos': # 正则余弦距离
        features[features > 0] = 1
        features = normalize(features, axis=1, norm='l1')
        # print(features.shape) # (30100, 24)
        dist = np.dot(features, features.T) # 乘法
        # print(dist.shape)     # (30100, 30100) 的对角阵

    inds = []
    for i in range(dist.shape[0]):
        ind = np.argpartition(dist[i, :], -(topk+1))[-(topk+1):] # 排序函数，按位置条件从小到大"排序"
        inds.append(ind)
    # print("inds:", len(inds), inds.shape) # 30100

    # f = open(fname, 'w')
    counter = 0
    A = np.zeros_like(dist)
    # print(A.shape) # (30100, 30100)
    # print("inds:", inds)
    src = []
    dst = []
    for i, v in enumerate(inds):
        mutual_knn = False
        for vv in v:
            if vv == i:
                pass
            else:
                if label[vv] != label[i]:
                    counter += 1
                src += [i]
                dst += [vv]
                # f.write('{} {}\n'.format(i, vv))
    # f.close()
    # print('error rate: {}'.format(counter / (num * topk)))
    return src, dst

class GATLayer(nn.Module):
    def __init__(self, g, in_dim , out_dim):
        super(GATLayer, self).__init__()
        self.g = g
        self.fc = nn.Linear(in_dim, out_dim, bias=False)
        self.attn_fc = nn.Linear(2*out_dim, 1, bias=False)
        self.reset_parameters()

    def reset_parameters(self):
        gain = nn.init.calculate_gain('relu')
        nn.init.xavier_normal_(self.fc.weight, gain=gain)
        nn.init.xavier_normal_(self.attn_fc.weight, gain=gain)

    def edge_attention(self, edges):
        z2 = torch.cat([edges.src['z'], edges.dst['z']], dim=1)
        a = self.attn_fc(z2)
        return {'e': F.leaky_relu(a)}

    def message_func(self,edges):
        return {'z': edges.src['z'], 'e': edges.data['e']}

    def reduce_func(self, nodes):
        alpha = F.softmax(nodes.mailbox['e'], dim=1) # 归一化每一条入边的注意力系数
        h = torch.sum(alpha * nodes.mailbox['z'], dim=1)
        return {'h':h}

    def forward(self, h):
        z = self.fc(h)
        self.g.ndata['z'] = z # 每个节点的特征
        self.g.apply_edges(self.edge_attention) # 为每一条边获得其注意力系数
        self.g.update_all(self.message_func, self.reduce_func)
        return self.g.ndata.pop('h')

class MultiHeadGATLayer(nn.Module):
    def __init__(self, g, in_dim , out_dim , num_heads=1, merge='cat'):
        super(MultiHeadGATLayer, self).__init__()
        self.heads = nn.ModuleList()
        for i in range(num_heads):
            self.heads.append(GATLayer(g, in_dim, out_dim))
        self.merge = merge


    def forward(self, h):
        head_out = [attn_head(h) for attn_head in self.heads]
        if self.merge=='cat':
            return torch.cat(head_out, dim=1)
        else:
            return torch.mean(torch.stack(head_out))

class GAT(nn.Module):
    def __init__(self, g, in_dim, hidden_dim , out_dim, num_heads):
        super(GAT, self).__init__()
        self.layer1 = MultiHeadGATLayer(g , in_dim, hidden_dim, num_heads)
        self.layer2 = MultiHeadGATLayer(g, hidden_dim*num_heads, 2, 1)
        # self.layer3 = Linear(out_dim,2)

    def forward(self, h):
        h = self.layer1(h)
        h = F.relu(h)
        h = self.layer2(h)
        h = torch.relu(h)
        # h = self.layer3(h)
        return h


class GAT_long(nn.Module): # net = GAT(graph, 39, hidden_dim=64, out_dim=8, num_heads=2).cuda()
    def __init__(self, g, in_dim, hidden_dim , out_dim, num_heads):
        super(GAT_long, self).__init__()
        self.layer1 = MultiHeadGATLayer(g , in_dim, hidden_dim, num_heads)
        self.layer2 = MultiHeadGATLayer(g, hidden_dim*num_heads, 64, 1)
        self.layer3 = MultiHeadGATLayer(g, 64, 32, 1)
        self.layer4 = MultiHeadGATLayer(g, 32, 8, 1)
        self.layer5 = MultiHeadGATLayer(g, 8, 2, 1)

    def forward(self, h):
        h = F.relu(self.layer1(h))
        h = F.relu(self.layer2(h))
        h = F.relu(self.layer3(h))
        h = F.relu(self.layer4(h))
        h = self.layer5(h)
        return h


def main():
    # 截取数据
    train_data = train_features
    train_label = train_labels
    print("Train:",train_data.shape) # [50, 24]

    # 源节点与目标节点建图
    src, dst = construct_graph(train_data, train_label, 'cos')
    u = torch.tensor(np.concatenate([src, dst])).cuda()
    v = torch.tensor(np.concatenate([dst, src])).cuda()
    graph = dgl.graph((u, v))
    print(graph)

    '''选取n个节点做loss计算'''
    # print(train_label[:40])
    # 400个节点，均匀分布
    labeled_nodes = t.tensor([i*10 for i in range(400)]).cuda() # 选出来要预测的节点，对应labels中的标签
    labels = t.tensor([train_labels[i*10] for i in range(400)]).cuda() 
    # labeled_nodes = t.tensor([35,40]).cuda() # 选出来要预测的节点，对应labels中的标签
    # labels = t.tensor([0, 1]).cuda()

    t.manual_seed(0) #固定初始化参数
    # net = GCN(39,128,2).cuda()
    # print("\nGCN结构:\n",net)
    
    net = GAT_long(graph, 39, hidden_dim=64, out_dim=8, num_heads=2).cuda()

    optimizer = t.optim.Adam(net.parameters(), lr=0.005)

    # 数据放到GPU上:训练数据放GPU，标签放CPU
    train_data = train_data.cuda()
    
    precision_max = 0
    for epoch in range(10000):
        # print(train.is_cuda)
        # outputs = net(graph, train_data) # 模型训练
        outputs = net(train_data) 
        # print("outputs.shape:",outputs.shape,outputs)

        logp = F.log_softmax(outputs, 1) #dim=1,对每一行进行softmax,当dim=0为按列
        # print("after_log_softmax:",logp)
        
        preds=t.argmax(F.softmax(outputs, 1),dim=1).cpu()
        # print("train_label:", train_label)
        # print("preds:", preds)
        # print(preds.is_cuda)
        correct_nodes = t.eq(preds,t.Tensor(train_label)).float().sum().item()
        # 对两个张量Tensor进行逐元素的比较，若相同位置的两个元素相同，则返回True；若不同，返回False。
        # print(len(preds), correct_nodes)
        precision = correct_nodes/len(preds)
        precision_max = max(precision_max, precision)
        # print("precision:" ,precision)
        
        # if precision > 0.76:
        #     torch.save(net.state_dict(), 'save_model/{}.pkl'.format(epoch))
        # compute loss for labeled nodes
        # print('\n',logp[labeled_nodes])
        loss = F.nll_loss(logp[labeled_nodes], labels.long())
        
        if loss < 0.001:
            break
        # compute gradient and do Adam step
        '''
        梯度归零 optimizer.zero_grad()
        反向传播计算得到每个参数的梯度值 loss.backward()
        通过梯度下降执行一步参数更新 optimizer.step()
        通过梯度下降执行一步参数更新 optimizer.step()
        '''
        optimizer.zero_grad()
        loss.backward()
        optimizer.step()
        
        print('\nEpoch %d | Loss: %.4f' % (epoch, loss.item()))        
        print('>>Fault: %d | Correct: %d | Precision: %f' % (len(preds)-correct_nodes,correct_nodes, precision))
    print("The best precision is ", precision_max)

if __name__ == "__main__":
    main()

