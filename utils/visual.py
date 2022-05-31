import pandas as pd
import numpy as np
import h5py
from responses import target
from sklearn.metrics.pairwise import cosine_similarity as cos
from sklearn.metrics import pairwise_distances as pair
from sklearn.preprocessing import normalize
import torch
import torch.nn as nn
torch.set_default_tensor_type(torch.FloatTensor)
from torch.autograd import Variable
from pandas.core.frame import DataFrame
from networkx import karate_club_graph,to_numpy_matrix
import matplotlib.pyplot as plt
import networkx as nx
import torch as t
import torch.nn.functional as F
import scipy.sparse as sp
from dgl.nn.pytorch import GraphConv
import dgl
from torch.nn import Linear
import openpyxl
 
from openpyxl import load_workbook


#########################################################
######################   数据加载  ######################
#########################################################

data_path = './our_data.xls'
# save_path = './visual.xlsx'

wb = openpyxl.Workbook()
ws = wb.active

data =  pd.read_excel(data_path, sheet_name=0)
data_select = []
data_target = []
k = 0
a, b = 0, 0
for row_i in data.index.values:
    feature_1 = data['最高气温均值'][row_i]
    feature_2 = data['最高气温方差'][row_i]
    feature_3 = data['平均气温均值'][row_i]
    if feature_3 > 40: # 异常气温
        continue
    
    if data['评审状态'][row_i] == "淘汰":
        # if a >= 2000: # 均衡采样
        #     continue
        data_target.append(0)
        a += 1
    elif data['评审状态'][row_i] == "续试" or data['评审状态'][row_i] == "续试1" :
        # if b >= 2000:
        #     continue
        data_target.append(1)
        b += 1
    else:
        continue
    feature_4 = data['平均气温方差'][row_i]
    feature_5 = data['最低气温均值'][row_i]
    feature_6 = data['最低气温方差'][row_i]
    feature_7 = data['温差均值'][row_i]
    feature_8 = data['温差方差'][row_i]
    feature_9 = data['地面气压均值'][row_i]
    feature_10 = data['地面气压方差'][row_i]
    feature_11 = data['相对湿度均值'][row_i]
    feature_12 = data['相对湿度方差'][row_i]
    feature_13 = data['降水量均值'][row_i]
    feature_14 = data['降水量方差'][row_i]
    feature_15 = data['最大风速均值'][row_i]
    feature_16 = data['最大风速方差'][row_i]
    feature_17 = data['平均风速均值'][row_i]
    feature_18 = data['平均风速方差'][row_i]
    feature_19 = data['风向角度均值'][row_i]
    feature_20 = data['风向角度方差'][row_i]
    feature_21 = data['日照时间均值'][row_i]
    feature_22 = data['日照时间方差'][row_i]
    feature_23 = data['风力等级均值'][row_i]
    feature_24 = data['风力等级方差'][row_i]
    feature_25 = data['大斑病'][row_i] 
    feature_26 = data['倒伏率'][row_i]
    feature_27 = data['倒折率'][row_i]
    feature_28 = data['灰斑病'][row_i]
    feature_29 = data['株高'][row_i]
    if feature_29 < 10:
        feature_29 *= 100
    feature_30 = data['穗位高'][row_i]
    feature_31 = data['空杆率'][row_i]
    feature_32 = data['生育期'][row_i]
    feature_33 = data['穗腐病'][row_i]
    feature_34 = data['百粒重'][row_i]
    feature_35 = data['穗长'][row_i] 
    feature_36 = data['秃尖长'][row_i]
    feature_37 = data['鲜穗果重'][row_i]
    feature_38 = data['亩产'][row_i]
    feature_39 = data['增减产'][row_i]

    data_select.append([feature_1, feature_2, feature_3, feature_4, feature_5, feature_6, feature_7, \
        feature_8, feature_9, feature_10, feature_11, feature_12, feature_13, feature_14, feature_15, \
        feature_16, feature_17, feature_18, feature_19, feature_20, feature_21, feature_22, feature_23,\
        feature_24, feature_25, feature_26, feature_27, feature_28, feature_29, feature_30, \
        feature_31, feature_32, feature_33, feature_34, feature_35, feature_36, feature_37,  \
        feature_38, feature_39])
    k += 1

print(k)
#########################################################
######################   数据处理  ######################
#########################################################
data_select = DataFrame(data_select)
print(data_select.shape)
print(len(data_select))

# print(len(data_select))
target = ["最高气温均值","最高气温方差","平均气温均值","平均气温方差",\
  "最低气温均值","最低气温方差","温差均值","温差方差","地面气压均值","地面气压方差","相对湿度均值",\
    "相对湿度方差","降水量均值","降水量方差","最大风速均值","最大风速方差","平均风速均值","平均风速方差",\
      "风向角度均值","风向角度方差","日照时间均值","日照时间方差","风力等级均值","风力等级方差","大斑病",\
        "倒伏率","倒折率","灰斑病","株高","穗位高","空杆率","生育期","穗腐病","百粒重","穗长",\
          "秃尖长","鲜穗果重","亩产","增减产","评审状态"]

# 特征做标准化
numeric_features = data_select.dtypes[data_select.dtypes != 'object'].index
data_select[numeric_features] = data_select[numeric_features].apply(
    lambda x: (x - x.mean()) / (x.std()))

# 特征做归一化
# numeric_features = data_select.dtypes[data_select.dtypes != 'object'].index
# data_select[numeric_features] = data_select[numeric_features].apply(
#     lambda x: (x - x.min()) / (x.max()-x.min()))

for j in range(39):
    xlsx_row = 2
    # print(data_select.iloc[:,j])
    ws.cell(row=1, column=j+1, value = target[j])
    for i in data_select.iloc[:,j]:
        ws.cell(row=xlsx_row, column=j+1, value = i)
        ws.cell(row=xlsx_row, column=40, value = data_target[xlsx_row-2])
        xlsx_row += 1
# for i range(len(data_select)):
#     ws.cell(row=1, column=j+1, value = target[j])
print(xlsx_row)

wb.save('./513_data.xlsx')
